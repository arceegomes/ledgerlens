#!/usr/bin/env python3
"""
Minimal Excel generation server for LedgerLens.
Uses openpyxl to fill the vendor template properly — preserving the Table,
data validations, dropdowns, and styles that SheetJS cannot round-trip.

Run alongside the Vite dev server:
    python excel_server.py

Listens on http://localhost:5175
Vite proxies /excel-api/* → here.
"""

import json
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import load_workbook
from io import BytesIO
import os

TEMPLATE = os.path.join(os.path.dirname(__file__), "public", "vendor-template.xlsx")

CORS = {
    "Access-Control-Allow-Origin":  "*",
    "Access-Control-Allow-Methods": "POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type",
}


class Handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        for k, v in CORS.items():
            self.send_header(k, v)
        self.end_headers()

    def do_POST(self):
        if self.path not in ("/generate-vendor-summary", "/generate-detail"):
            self.send_response(404)
            self.end_headers()
            return

        length = int(self.headers.get("Content-Length", 0))
        body   = json.loads(self.rfile.read(length))
        rows   = body.get("rows", [])

        wb = load_workbook(TEMPLATE)
        ws = wb["Sheet1"]

        # Clear any sample rows from the template (rows 2+)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # Write data rows
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val if val != "" else None)

        # Expand the Table1 ref to cover all data rows
        last_row = max(len(rows) + 1, 2)   # at least header + 1 row
        ws.tables["Table1"].ref = f"A1:N{last_row}"

        buf = BytesIO()
        wb.save(buf)
        xlsx = buf.getvalue()

        self.send_response(200)
        self.send_header("Content-Type",
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Length", str(len(xlsx)))
        for k, v in CORS.items():
            self.send_header(k, v)
        self.end_headers()
        self.wfile.write(xlsx)

    def log_message(self, fmt, *args):
        print(f"[excel] {fmt % args}")


if __name__ == "__main__":
    port = 5175
    print(f"LedgerLens Excel server -> http://localhost:{port}")
    print(f"Template: {TEMPLATE}")
    HTTPServer(("localhost", port), Handler).serve_forever()
